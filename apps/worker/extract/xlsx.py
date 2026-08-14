from openpyxl import load_workbook


def extract_text(path: str) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append("\t".join(cells))
    return "\n".join(lines)
